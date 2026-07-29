"""Candidate registration ingestion.

Mirrors the ExaMetrics "fresh import" round-trip (``GET /fresh-import/template``
→ fill → ``POST /fresh-import/upload``) for LAZEIMS, minus the marks: LAZEIMS
collects marks through its own attendance-gated pipeline, so a register only
carries candidate identity plus which subjects each candidate sits.

The flow is deliberately **stateless**:

    template  →  preview (parse + validate, persists nothing)  →  bulk (apply)

Parsed rows round-trip through the client, so there is no staging table and no
migration. The same ``preview`` → ``bulk`` funnel serves all three ingestion
doors — Excel/CSV upload, and rows extracted from a registration PDF/ZIP by
``POST /exams/{id}/registrations/extract``.

Register layout (generated and accepted):

    row 1    CANDIDATE REGISTER
    row 2    Exam: <name>
    row 3    Centre: <centre_number> — <school name>
    row 4    (instructions)
    row 5    CNO | FULL NAME | SEX | <subject code> | <subject code> | …
    row 6+   data

The header row is *located by scanning* for a first cell of ``CNO``, so
hand-built files with a different preamble still import.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from lazeims_common.enums import RejectionCode, Sex
from lazeims_common.errors import ValidationError

from ..models.exam import Exam, ExamSchool, ExamStudent, ExamStudentSubject, ExamSubject
from ..models.registry import School, Subject

SHEET_NAME = "REGISTER"
ID_HEADER = "CNO"
NAME_HEADER = "FULL NAME"
SEX_HEADER = "SEX"
FIXED_HEADERS = (ID_HEADER, NAME_HEADER, SEX_HEADER)

MAX_TEMPLATE_ROWS = 2000
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# Row status values returned by preview / consumed by apply.
STATUS_OK = "OK"
STATUS_DUPLICATE = "DUPLICATE"
STATUS_ERROR = "ERROR"

_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_META_FONT = Font(bold=True, size=11)


# ─────────────────────────────── helpers ────────────────────────────────


def _norm(value: Any) -> str:
    """Trim and collapse internal whitespace; ``None`` becomes ``""``."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def split_full_name(raw: str) -> tuple[str, str | None, str]:
    """Split a display name into ``(first, middle, surname)``.

    Accepts both NECTA-style ``"SURNAME, First Middle"`` and plain
    ``"First Middle Surname"``. Returns empty strings when unsplittable so the
    caller can raise a per-row error rather than guessing.
    """
    name = _norm(raw)
    if not name:
        return "", None, ""

    if "," in name:
        surname_part, _, rest = name.partition(",")
        surname = _norm(surname_part)
        given = _norm(rest).split(" ") if _norm(rest) else []
        if not given:
            return surname, None, surname
        first = given[0]
        middle = " ".join(given[1:]) or None
        return first, middle, surname

    parts = name.split(" ")
    if len(parts) == 1:
        return parts[0], None, parts[0]
    if len(parts) == 2:
        return parts[0], None, parts[1]
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def _parse_sex(raw: str) -> str | None:
    token = _norm(raw).upper()
    if token in ("M", "MALE", "ME"):
        return Sex.M.value
    if token in ("F", "FEMALE", "KE"):
        return Sex.F.value
    return None


def _is_ticked(value: Any) -> bool:
    """A subject cell counts as 'sits this subject' when non-empty and not a
    negative marker. Accepts X / x / 1 / TRUE / YES / a mark value."""
    token = _norm(value).upper()
    if not token:
        return False
    return token not in ("0", "-", "N", "NO", "FALSE", "NA", "N/A")


async def _exam_subject_index(db: AsyncSession, exam_id) -> dict[str, int]:
    """Map subject code (upper) → exam_subject id for this exam."""
    rows = (
        await db.execute(
            select(Subject.code, ExamSubject.id)
            .join(ExamSubject, ExamSubject.subject_id == Subject.id)
            .where(ExamSubject.exam_id == exam_id)
            .order_by(Subject.code)
        )
    ).all()
    return {str(code).upper(): es_id for code, es_id in rows}


async def _require_enrolled_school(db: AsyncSession, exam_id, school_id: int) -> School:
    school = await db.get(School, school_id)
    if school is None:
        raise ValidationError(
            RejectionCode.INVALID_NATURAL_KEY,
            "School not found.",
            {"school_id": school_id},
        )
    enrolled = await db.scalar(
        select(func.count())
        .select_from(ExamSchool)
        .where(ExamSchool.exam_id == exam_id, ExamSchool.school_id == school_id)
    )
    if not enrolled:
        raise ValidationError(
            RejectionCode.OUTSIDE_STATION_SCOPE,
            "That school is not enrolled in this exam. Enrol it first.",
            {"school_id": school_id, "exam_id": str(exam_id)},
        )
    return school


# ───────────────────────── template generation ──────────────────────────


async def generate_register_template(
    db: AsyncSession,
    *,
    exam: Exam,
    school: School | None,
    rows: int = 400,
) -> bytes:
    """Build the candidate register workbook for an exam (optionally scoped to
    one school), pre-headed with the exam's subject codes."""
    rows = max(1, min(rows, MAX_TEMPLATE_ROWS))
    codes = sorted((await _exam_subject_index(db, exam.id)).keys())

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.cell(row=1, column=1, value="CANDIDATE REGISTER").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Exam: {exam.name}").font = _META_FONT
    centre = (
        f"Centre: {school.centre_number} — {school.name}"
        if school is not None
        else "Centre: (fill one school per sheet)"
    )
    ws.cell(row=3, column=1, value=centre).font = _META_FONT
    ws.cell(
        row=4,
        column=1,
        value=(
            "Fill CNO, FULL NAME (First Middle Surname or SURNAME, First Middle) and "
            "SEX (M/F). Put X under every subject the candidate sits. Do not rename or "
            "reorder the header row."
        ),
    ).font = Font(italic=True, size=9)

    header = [*FIXED_HEADERS, *codes]
    for col, label in enumerate(header, start=1):
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 7
    for idx in range(len(FIXED_HEADERS) + 1, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 8

    ws.freeze_panes = "A6"
    # Reserve the data band so the sheet opens at a usable size.
    ws.cell(row=5 + rows, column=1, value=None)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────── parsing ────────────────────────────────


def _locate_header(grid: list[list[Any]]) -> int:
    """Return the 0-based index of the header row, or raise."""
    for idx, row in enumerate(grid):
        if row and _norm(row[0]).upper() == ID_HEADER:
            return idx
    raise ValidationError(
        RejectionCode.CONFIGURATION_MISMATCH,
        f"Could not find the header row. Row 5 must start with '{ID_HEADER}'.",
        {"expected_headers": list(FIXED_HEADERS)},
    )


def _grid_from_xlsx(file_bytes: bytes) -> list[list[Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _grid_from_csv(file_bytes: bytes) -> list[list[Any]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return [list(row) for row in csv.reader(io.StringIO(text))]


def parse_register_file(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse an uploaded register into raw row dicts.

    Returns ``[{row_no, student_id, full_name, sex, subject_codes[]}]`` — no
    validation beyond structural parsing.
    """
    if not file_bytes:
        raise ValidationError(RejectionCode.CONFIGURATION_MISMATCH, "The file is empty.", {})
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise ValidationError(
            RejectionCode.CONFIGURATION_MISMATCH,
            "Register must not exceed 10 MB.",
            {"size": len(file_bytes)},
        )

    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        grid = _grid_from_xlsx(file_bytes)
    elif lower.endswith(".csv"):
        grid = _grid_from_csv(file_bytes)
    else:
        raise ValidationError(
            RejectionCode.CONFIGURATION_MISMATCH,
            "Register must be .xlsx, .xlsm or .csv",
            {"filename": filename},
        )

    header_idx = _locate_header(grid)
    header = [_norm(c).upper() for c in grid[header_idx]]
    subject_cols = [
        (idx, code)
        for idx, code in enumerate(header)
        if idx >= len(FIXED_HEADERS) and code
    ]

    out: list[dict[str, Any]] = []
    for offset, raw in enumerate(grid[header_idx + 1 :], start=header_idx + 2):
        if not raw or not any(_norm(c) for c in raw):
            continue
        student_id = _norm(raw[0] if len(raw) > 0 else "")
        full_name = _norm(raw[1] if len(raw) > 1 else "")
        sex = _norm(raw[2] if len(raw) > 2 else "")
        if not student_id and not full_name:
            continue
        codes = [
            code
            for idx, code in subject_cols
            if idx < len(raw) and _is_ticked(raw[idx])
        ]
        out.append(
            {
                "row_no": offset,
                "student_id": student_id,
                "full_name": full_name,
                "sex": sex,
                "subject_codes": codes,
            }
        )
    return out


# ────────────────────────────── validation ──────────────────────────────


async def validate_rows(
    db: AsyncSession,
    *,
    exam: Exam,
    school_id: int,
    raw_rows: Iterable[dict[str, Any]],
) -> dict[str, Any]:
    """Validate raw rows against the exam's subjects and existing registrations.

    Persists nothing. Returns a preview payload whose ``rows`` can be posted
    straight back to the bulk-apply endpoint.
    """
    school = await _require_enrolled_school(db, exam.id, school_id)
    code_to_exam_subject = await _exam_subject_index(db, exam.id)

    existing_ids = set(
        (
            await db.execute(
                select(ExamStudent.student_id).where(ExamStudent.exam_id == exam.id)
            )
        ).scalars().all()
    )

    seen: set[str] = set()
    rows: list[dict[str, Any]] = []
    counts = {STATUS_OK: 0, STATUS_DUPLICATE: 0, STATUS_ERROR: 0}

    for raw in raw_rows:
        row_no = raw.get("row_no")
        student_id = _norm(raw.get("student_id"))
        errors: list[str] = []

        # Identity may arrive pre-split (PDF extract) or as one display name.
        first = _norm(raw.get("first_name"))
        middle = _norm(raw.get("middle_name")) or None
        surname = _norm(raw.get("surname"))
        if not (first and surname):
            first, middle, surname = split_full_name(raw.get("full_name") or "")

        sex = _parse_sex(raw.get("sex") or "")

        if not student_id:
            errors.append("Missing candidate number (CNO).")
        elif len(student_id) > 60:
            errors.append("Candidate number exceeds 60 characters.")
        if not (first and surname):
            errors.append("Name must contain at least a first name and a surname.")
        if sex is None:
            errors.append("Sex must be M or F.")

        codes = [
            _norm(c).upper()
            for c in (raw.get("subject_codes") or [])
            if _norm(c)
        ]
        unknown = [c for c in codes if c not in code_to_exam_subject]
        if unknown:
            errors.append(
                "Subject(s) not offered in this exam: " + ", ".join(sorted(set(unknown)))
            )
        if not codes and not errors:
            errors.append("No subjects ticked for this candidate.")

        status = STATUS_OK
        if errors:
            status = STATUS_ERROR
        elif student_id in seen:
            status = STATUS_ERROR
            errors.append("Duplicate candidate number within this file.")
        elif student_id in existing_ids:
            status = STATUS_DUPLICATE
            errors.append("Already registered in this exam — will be skipped.")

        if student_id:
            seen.add(student_id)
        counts[status] += 1

        rows.append(
            {
                "row_no": row_no,
                "student_id": student_id,
                "first_name": first,
                "middle_name": middle,
                "surname": surname,
                "sex": sex,
                "subject_codes": codes,
                "exam_subject_ids": [
                    code_to_exam_subject[c] for c in codes if c in code_to_exam_subject
                ],
                "status": status,
                "errors": errors,
            }
        )

    return {
        "school_id": school_id,
        "centre_number": school.centre_number,
        "school_name": school.name,
        "row_count": len(rows),
        "ok_count": counts[STATUS_OK],
        "duplicate_count": counts[STATUS_DUPLICATE],
        "error_count": counts[STATUS_ERROR],
        "known_subject_codes": sorted(code_to_exam_subject.keys()),
        "rows": rows,
    }


# ──────────────────────────────── apply ─────────────────────────────────


async def apply_rows(
    db: AsyncSession,
    *,
    exam: Exam,
    school_id: int,
    rows: Iterable[dict[str, Any]],
) -> dict[str, Any]:
    """Register the supplied candidates. Idempotent: candidates whose
    ``student_id`` already exists in the exam are skipped, not failed.

    Rows are re-validated server-side — a client cannot bypass validation by
    editing the preview payload.
    """
    preview = await validate_rows(db, exam=exam, school_id=school_id, raw_rows=rows)

    created = 0
    skipped = 0
    failed: list[dict[str, Any]] = []
    subject_links = 0

    for row in preview["rows"]:
        if row["status"] == STATUS_DUPLICATE:
            skipped += 1
            continue
        if row["status"] == STATUS_ERROR:
            failed.append(
                {
                    "row_no": row["row_no"],
                    "student_id": row["student_id"],
                    "errors": row["errors"],
                }
            )
            continue

        student = ExamStudent(
            student_id=row["student_id"],
            exam_id=exam.id,
            school_id=school_id,
            first_name=row["first_name"],
            middle_name=row["middle_name"],
            surname=row["surname"],
            sex=row["sex"],
        )
        db.add(student)
        await db.flush()

        for exam_subject_id in dict.fromkeys(row["exam_subject_ids"]):
            db.add(
                ExamStudentSubject(
                    exam_student_id=student.id, exam_subject_id=exam_subject_id
                )
            )
            subject_links += 1
        created += 1

    await db.flush()

    return {
        "school_id": school_id,
        "centre_number": preview["centre_number"],
        "created": created,
        "skipped": skipped,
        "failed": len(failed),
        "subject_registrations": subject_links,
        "errors": failed,
    }


# ──────────────────────────────── stats ─────────────────────────────────


def rows_from_extract(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Flatten an ExaMetrics extraction envelope into per-centre groups.

    ExaMetrics' ``POST /integration/registrations/extract`` returns one document
    per PDF — a ZIP of registers therefore yields many centres in one upload.
    Its rows are already normalised (``student_id``, ``first_name``,
    ``middle_name``, ``surname``, ``sex``, ``combination``, ``subject_codes``),
    so nothing needs guessing from column headers.

    Returns ``[{centre_number, school_name, exam_level, exam_year, source,
    parse_error, rows[]}]`` — one entry per document, preserving failures so the
    caller can report which register in a ZIP could not be read.
    """
    groups: list[dict[str, Any]] = []

    for document in payload.get("documents") or []:
        source = document.get("source")

        # A document that failed to parse carries `error` instead of `school`.
        if document.get("error"):
            groups.append(
                {
                    "source": source,
                    "centre_number": None,
                    "school_name": None,
                    "exam_level": None,
                    "exam_year": None,
                    "parse_error": str(document["error"]),
                    "rows": [],
                }
            )
            continue

        school = document.get("school") or {}
        rows: list[dict[str, Any]] = []
        for raw in document.get("rows") or []:
            rows.append(
                {
                    "student_id": _norm(raw.get("student_id")),
                    "first_name": _norm(raw.get("first_name")),
                    "middle_name": _norm(raw.get("middle_name")) or None,
                    "surname": _norm(raw.get("surname")),
                    "sex": _norm(raw.get("sex")),
                    "combination": _norm(raw.get("combination")) or None,
                    "subject_codes": [
                        _norm(code).upper()
                        for code in (raw.get("subject_codes") or [])
                        if _norm(code)
                    ],
                }
            )

        groups.append(
            {
                "source": source,
                "centre_number": _norm(school.get("centre_number")).upper() or None,
                "school_name": _norm(school.get("school_name")) or None,
                "exam_level": _norm(school.get("exam_level")) or None,
                "exam_year": _norm(school.get("exam_year")) or None,
                "parse_error": None,
                "rows": rows,
            }
        )

    return groups


async def resolve_centres(
    db: AsyncSession, *, exam_id, centre_numbers: Iterable[str]
) -> dict[str, dict[str, Any]]:
    """Map centre number → registry school and whether it is enrolled in the exam.

    A register names its centre; LAZEIMS keys candidates by ``school_id``. This
    is the join, done once for a whole batch.
    """
    wanted = {c.upper() for c in centre_numbers if c}
    if not wanted:
        return {}

    rows = (
        await db.execute(
            select(School.id, School.centre_number, School.name).where(
                func.upper(School.centre_number).in_(wanted)
            )
        )
    ).all()

    enrolled = set(
        (
            await db.execute(
                select(ExamSchool.school_id).where(ExamSchool.exam_id == exam_id)
            )
        ).scalars().all()
    )

    return {
        str(centre).upper(): {
            "school_id": school_id,
            "centre_number": centre,
            "school_name": name,
            "enrolled": school_id in enrolled,
        }
        for school_id, centre, name in rows
    }


async def registration_stats(db: AsyncSession, *, exam_id) -> dict[str, Any]:
    """Per-school candidate counts, so the UI can show registration progress
    without pulling every candidate."""
    rows = (
        await db.execute(
            select(
                School.id,
                School.centre_number,
                School.name,
                func.count(ExamStudent.id),
            )
            .select_from(ExamSchool)
            .join(School, School.id == ExamSchool.school_id)
            .outerjoin(
                ExamStudent,
                (ExamStudent.school_id == ExamSchool.school_id)
                & (ExamStudent.exam_id == exam_id),
            )
            .where(ExamSchool.exam_id == exam_id)
            .group_by(School.id, School.centre_number, School.name)
            .order_by(School.centre_number)
        )
    ).all()

    schools = [
        {
            "school_id": sid,
            "centre_number": centre,
            "school_name": name,
            "candidate_count": count or 0,
        }
        for sid, centre, name, count in rows
    ]
    with_candidates = sum(1 for s in schools if s["candidate_count"] > 0)
    return {
        "school_count": len(schools),
        "schools_with_candidates": with_candidates,
        "schools_pending": len(schools) - with_candidates,
        "candidate_total": sum(s["candidate_count"] for s in schools),
        "schools": schools,
    }
