"""Controlled Excel collection service (Guide §2.7, delivery §12).

* Workbook generation is scoped to ONE school (only when
  ``can_download_template`` is granted), one subject set, its papers. v1 supports
  TOTAL_MARKS exams (one editable total per student-paper); the workbook embeds
  its ``workbook_id`` + ``configuration_hash`` in a protected meta sheet.
* Import streams the workbook, stages every row, and dry-runs the SAME
  ``lazeims_common`` validation used by online/station — no weaker Excel path.
* A separate confirm step applies accepted rows through the same attendance +
  marks services (EXCEL writer channel) with idempotency; replay applies once.
"""

from __future__ import annotations

import hashlib
import io
import uuid
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from lazeims_common.enums import FillingMode, PaperType, RejectionCode, WriterMode
from lazeims_common.errors import ValidationError
from lazeims_common.hashing import sha256_prefixed
from lazeims_common.validation.marks import validate_marks_submission

from ..models.exam import Exam, ExamStudent, ExamStudentSubject, ExamSubject
from ..models.excel import ExcelImportBatch, ExcelImportRow, ExcelWorkbook
from ..models.registry import School, Subject
from .attendance import upsert_attendance
from .marks_apply import apply_student_paper_marks, build_paper_config
from .scope_assignment import is_scope_finalized, resolve_student_scope

META_SHEET = "_meta"
HEADER = ["student_id", "name", "present", "total"]


def _applicable_papers(es: ExamSubject) -> list[PaperType]:
    papers = [PaperType.THEORY1]
    if es.has_theory2:
        papers.append(PaperType.THEORY2)
    if es.has_practical:
        papers.append(PaperType.PRACTICAL)
    return papers


def _sheet_name(subject_code: str, paper: PaperType) -> str:
    return f"{subject_code}-{paper.value}"[:31]


async def _build_workbook_bytes(
    db: AsyncSession, *, exam: Exam, school: School, exam_subjects: list,
    workbook_id: str, config_hash: str,
) -> tuple[bytes, dict]:
    wb = Workbook()
    wb.remove(wb.active)
    papers_used: set[str] = set()
    subjects_used: set[str] = set()

    for es in exam_subjects:
        subject = await db.get(Subject, es.subject_id)
        subjects_used.add(subject.code)
        rows = (
            await db.execute(
                select(ExamStudent)
                .join(ExamStudentSubject, ExamStudentSubject.exam_student_id == ExamStudent.id)
                .where(ExamStudent.exam_id == exam.id, ExamStudent.school_id == school.id,
                       ExamStudentSubject.exam_subject_id == es.id)
                .order_by(ExamStudent.student_id)
            )
        ).scalars().all()
        for paper in _applicable_papers(es):
            papers_used.add(paper.value)
            ws = wb.create_sheet(_sheet_name(subject.code, paper))
            ws.append(HEADER)
            for st in rows:
                ws.append([st.student_id, f"{st.first_name} {st.surname}", "Y", None])

    meta = wb.create_sheet(META_SHEET, 0)
    meta.append(["workbook_id", workbook_id])
    meta.append(["exam_id", str(exam.id)])
    meta.append(["centre_number", school.centre_number])
    meta.append(["configuration_hash", config_hash])
    meta.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    scope = {"subjects": sorted(subjects_used), "papers": sorted(papers_used)}
    return buf.getvalue(), scope


async def _load_exam_subjects(db, exam, subject_codes):
    return (
        await db.execute(
            select(ExamSubject).join(Subject, Subject.id == ExamSubject.subject_id)
            .where(ExamSubject.exam_id == exam.id, Subject.code.in_(subject_codes))
        )
    ).scalars().all()


async def generate_workbook(
    db: AsyncSession, *, exam: Exam, school: School, subject_codes: list[str], generated_by: int | None,
) -> tuple[ExcelWorkbook, bytes]:
    if not school.can_download_template:
        raise ValidationError(
            RejectionCode.OUTSIDE_STATION_SCOPE,
            "This school is not granted Excel template download (can_download_template is off).",
            {"centre_number": school.centre_number},
        )
    if (exam.settings or {}).get("filling_mode", "TOTAL_MARKS") != "TOTAL_MARKS":
        raise ValidationError(
            RejectionCode.CONFIGURATION_MISMATCH,
            "Excel collection currently supports total-marks exams only.",
        )
    exam_subjects = await _load_exam_subjects(db, exam, subject_codes)
    if not exam_subjects:
        raise ValidationError(RejectionCode.NOT_REGISTERED, "No matching subjects offered in this exam.")

    subject_scope = {"subjects": sorted(subject_codes),
                     "papers": sorted({p.value for es in exam_subjects for p in _applicable_papers(es)})}
    config_hash = sha256_prefixed({"exam_id": str(exam.id), "centre": school.centre_number, "scope": subject_scope})
    workbook_id = "wbk_" + hashlib.sha256(
        f"{str(exam.id)}:{school.centre_number}:{uuid.uuid4().hex}".encode()).hexdigest()[:20]

    content, scope = await _build_workbook_bytes(
        db, exam=exam, school=school, exam_subjects=exam_subjects,
        workbook_id=workbook_id, config_hash=config_hash)

    row = ExcelWorkbook(
        workbook_id=workbook_id, exam_id=exam.id, school_id=school.id,
        subject_scope=scope, configuration_hash=config_hash,
        generated_by=generated_by, generated_at=datetime.now(timezone.utc), status="ISSUED",
    )
    db.add(row)
    await db.flush()
    return row, content


async def render_workbook_bytes(db: AsyncSession, workbook: ExcelWorkbook) -> bytes:
    """Rebuild the SAME workbook's bytes (same id/hash) for re-download."""
    exam = await db.get(Exam, workbook.exam_id)
    school = await db.get(School, workbook.school_id)
    exam_subjects = await _load_exam_subjects(db, exam, workbook.subject_scope["subjects"])
    content, _scope = await _build_workbook_bytes(
        db, exam=exam, school=school, exam_subjects=exam_subjects,
        workbook_id=workbook.workbook_id, config_hash=workbook.configuration_hash)
    return content


def _read_meta(wb) -> dict:
    meta = {}
    if META_SHEET in wb.sheetnames:
        for r in wb[META_SHEET].iter_rows(values_only=True):
            if r and r[0]:
                meta[r[0]] = r[1]
    return meta


async def parse_and_stage(
    db: AsyncSession, *, workbook: ExcelWorkbook, file_bytes: bytes, channel: str,
    idempotency_key: str, imported_by: int | None,
) -> ExcelImportBatch:
    """Stream-parse a returned workbook, stage every row, and dry-run validate."""
    # Idempotent: same key -> return the existing batch.
    existing = (
        await db.execute(select(ExcelImportBatch).where(ExcelImportBatch.idempotency_key == idempotency_key))
    ).scalar_one_or_none()
    if existing is not None:
        return existing

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    meta = _read_meta(wb)
    if meta.get("workbook_id") != workbook.workbook_id:
        raise ValidationError(
            RejectionCode.CONFIGURATION_MISMATCH,
            "Uploaded workbook does not match the issued workbook_id.",
            {"expected": workbook.workbook_id, "actual": meta.get("workbook_id")},
        )

    exam = await db.get(Exam, workbook.exam_id)
    import_id = "imp_" + uuid.uuid4().hex[:20]
    batch = ExcelImportBatch(
        import_id=import_id, workbook_id=workbook.workbook_id, imported_by=imported_by,
        channel=channel, idempotency_key=idempotency_key, status="STAGED",
    )
    db.add(batch)
    await db.flush()

    row_no = 0
    accepted = rejected = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == META_SHEET:
            continue
        try:
            subject_code, paper_value = sheet_name.rsplit("-", 1)
            paper = PaperType(paper_value)
        except (ValueError, KeyError):
            continue
        es = (
            await db.execute(select(ExamSubject).join(Subject, Subject.id == ExamSubject.subject_id)
                             .where(ExamSubject.exam_id == exam.id, Subject.code == subject_code))
        ).scalar_one_or_none()
        ws = wb[sheet_name]
        header_seen = False
        for raw in ws.iter_rows(values_only=True):
            if not header_seen:
                header_seen = True
                continue
            if raw is None or all(c is None for c in raw):
                continue
            student_id = str(raw[0]).strip() if raw[0] is not None else None
            present_cell = (str(raw[2]).strip().upper() if len(raw) > 2 and raw[2] is not None else "Y")
            total_cell = raw[3] if len(raw) > 3 else None
            is_present = present_cell in ("Y", "YES", "TRUE", "1", "PRESENT")
            row_no += 1
            payload = {"student_id": student_id, "subject_code": subject_code,
                       "paper_type": paper.value, "is_present": is_present,
                       "total": (None if total_cell is None or str(total_cell).strip() == "" else str(total_cell))}
            code = msg = None
            try:
                if es is None:
                    raise ValidationError(RejectionCode.NOT_REGISTERED, f"Subject {subject_code} not offered.")
                paper_cfg = await build_paper_config(db, es, paper)
                validate_marks_submission(
                    mode=FillingMode.TOTAL_MARKS, is_present=is_present,
                    has_attendance_transcription=True, paper=paper_cfg,
                    total_marks_obtained=(None if payload["total"] is None else Decimal(payload["total"])),
                    item_marks=None,
                )
            except ValidationError as exc:
                code, msg = exc.code.value, exc.message
            except Exception as exc:  # noqa: BLE001 - malformed cell
                code, msg = RejectionCode.BLANK_MARK_NOT_ALLOWED.value, f"Unparseable value: {exc}"

            status = "OK" if code is None else "ERROR"
            accepted += status == "OK"
            rejected += status == "ERROR"
            db.add(ExcelImportRow(
                import_id=import_id, row_no=row_no, student_id=student_id,
                subject_code=subject_code, paper_type=paper.value, payload=payload,
                status=status, error_code=code, error_message=msg,
            ))
    batch.row_count = row_no
    batch.accepted_count = accepted
    batch.rejected_count = rejected
    workbook.status = "RETURNED"
    await db.flush()
    return batch


async def confirm_import(db: AsyncSession, *, batch: ExcelImportBatch, confirmed_by: int | None) -> dict:
    """Apply staged OK rows through the same attendance + marks services (EXCEL
    writer). Idempotent: a second confirm is a no-op returning the same result."""
    workbook = await db.get(ExcelWorkbook, batch.workbook_id)
    exam = await db.get(Exam, workbook.exam_id)

    if batch.status == "APPLIED":
        return {"applied": True, "already": True,
                "accepted": batch.accepted_count, "rejected": batch.rejected_count}

    rows = (
        await db.execute(select(ExcelImportRow).where(
            ExcelImportRow.import_id == batch.import_id, ExcelImportRow.status == "OK"))
    ).scalars().all()

    applied = 0
    for row in rows:
        paper = PaperType(row.paper_type)
        scope = await resolve_student_scope(
            db, exam_id=exam.id, student_id=row.student_id,
            exam_subject_id=(await _exam_subject_id(db, exam.id, row.subject_code)))
        if await is_scope_finalized(db, exam_id=exam.id, school_id=scope.student.school_id,
                                    exam_subject_id=scope.exam_subject.id, paper_type=paper):
            continue
        # attendance (from the 'present' column) then marks — same services as online.
        from lazeims_common.enums import AttendanceSource
        await upsert_attendance(db, exam_student_subject_id=scope.exam_student_subject.id,
                                paper_type=paper, is_present=row.payload["is_present"],
                                source=AttendanceSource.INVIGILATOR_ISAL_TRANSCRIPTION, actor_id=confirmed_by)
        await apply_student_paper_marks(
            db, scope=scope, paper_type=paper, mode=FillingMode.TOTAL_MARKS,
            total_marks_obtained=(None if row.payload["total"] is None else Decimal(row.payload["total"])),
            items=None, actor_id=confirmed_by)
        row.status = "APPLIED"
        applied += 1

    batch.status = "APPLIED"
    batch.imported_at = datetime.now(timezone.utc)
    workbook.status = "IMPORTED"
    await db.flush()
    return {"applied": True, "already": False, "rows_applied": applied,
            "accepted": batch.accepted_count, "rejected": batch.rejected_count}


async def _exam_subject_id(db, exam_id, subject_code) -> int:
    es = (
        await db.execute(select(ExamSubject).join(Subject, Subject.id == ExamSubject.subject_id)
                         .where(ExamSubject.exam_id == exam_id, Subject.code == subject_code))
    ).scalar_one_or_none()
    if es is None:
        raise ValidationError(RejectionCode.NOT_REGISTERED, f"Subject {subject_code} not offered.")
    return es.id
