from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import load_workbook

from app.db import get_sessionmaker
from tests.conftest import login

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _admin(client):
    csrf = await login(client, "superadmin", "adminpass123")
    return {"X-CSRF-Token": csrf}


async def _reg(client, path, json, h):
    r = await client.post(f"/api/v1/registry/{path}", json=json, headers=h)
    assert r.status_code == 201, r.text
    return r.json()


async def _post(client, path, json, h, expect=201):
    r = await client.post(f"/api/v1/{path}", json=json, headers=h)
    assert r.status_code == expect, r.text
    return r.json()


async def _setup(client, h, *, can_download=True, writer="EXCEL"):
    school = (await _reg(client, "schools", {"centre_number": f"SCH-{uuid.uuid4().hex[:5]}",
              "name": "S", "can_download_template": can_download}, h))
    school_id, centre = school["id"], school["centre_number"]
    subject = (await _reg(client, "subjects", {"code": "011", "name": "H"}, h))["id"]
    exam_id = (await _post(client, "exams", {"name": "E", "exam_code": f"EX-{uuid.uuid4().hex[:6]}", "level_id": 1}, h))["id"]
    await _post(client, f"exams/{exam_id}/schools", {"school_id": school_id}, h)
    es = (await _post(client, f"exams/{exam_id}/subjects", {"subject_id": subject, "total_marks_theory1": 100}, h))["id"]
    for sid in ("S-1", "S-2"):
        await _post(client, f"exams/{exam_id}/students",
                    {"student_id": sid, "school_id": school_id, "first_name": sid, "surname": "X",
                     "sex": "M", "subject_ids": [es]}, h)
    await client.put(f"/api/v1/exams/{exam_id}/writer-assignments", json={
        "school_id": school_id, "exam_subject_id": es, "paper_type": "THEORY1", "writer_mode": writer}, headers=h)
    await _post(client, f"exams/{exam_id}/transitions", {"target_phase": "ENTRY_OPEN"}, h, expect=200)
    return {"exam_id": exam_id, "school_id": school_id, "centre": centre, "es": es}


def _fill(xbytes, edits):
    """edits: {student_id: (present, total)} on sheet 011-THEORY1."""
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["011-THEORY1"]
    for row in ws.iter_rows(min_row=2):
        sid = row[0].value
        if sid in edits:
            present, total = edits[sid]
            row[2].value = present
            row[3].value = total
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def _make_workbook_bytes(client, h, ctx):
    wb = await _post(client, f"exams/{ctx['exam_id']}/excel/workbooks",
                     {"school_id": ctx["school_id"], "subjects": ["011"]}, h, expect=200)
    dl = await client.get(f"/api/v1/excel/workbooks/{wb['workbook_id']}/download", headers=h)
    assert dl.status_code == 200 and dl.headers["content-type"] == XLSX
    return wb["workbook_id"], dl.content


# ---------- grant gate ----------

async def test_workbook_refused_without_grant(client):
    h = await _admin(client)
    ctx = await _setup(client, h, can_download=False)
    r = await client.post(f"/api/v1/exams/{ctx['exam_id']}/excel/workbooks",
                          json={"school_id": ctx["school_id"], "subjects": ["011"]}, headers=h)
    assert r.status_code == 422 and r.json()["detail"]["code"] == "OUTSIDE_STATION_SCOPE"


async def test_workbook_generated_with_grant(client):
    h = await _admin(client)
    ctx = await _setup(client, h)
    wid, content = await _make_workbook_bytes(client, h, ctx)
    wb = load_workbook(io.BytesIO(content))
    assert "011-THEORY1" in wb.sheetnames


# ---------- validation parity ----------

async def test_import_preview_flags_invalid_rows(client):
    h = await _admin(client)
    ctx = await _setup(client, h)
    wid, content = await _make_workbook_bytes(client, h, ctx)
    # S-1 valid (60), S-2 over max (200)
    filled = _fill(content, {"S-1": ("Y", 60), "S-2": ("Y", 200)})
    r = await client.post(f"/api/v1/excel/workbooks/{wid}/imports",
                          files={"file": ("wb.xlsx", filled, XLSX)},
                          headers={**h, "Idempotency-Key": uuid.uuid4().hex})
    body = r.json()
    assert body["accepted_count"] == 1 and body["rejected_count"] == 1
    errs = (await client.get(f"/api/v1/excel/imports/{body['import_id']}/errors", headers=h)).json()["errors"]
    assert errs[0]["code"] == "MARK_OUT_OF_RANGE" and errs[0]["student_id"] == "S-2"


async def test_absent_with_marks_flagged(client):
    h = await _admin(client)
    ctx = await _setup(client, h)
    wid, content = await _make_workbook_bytes(client, h, ctx)
    filled = _fill(content, {"S-1": ("N", 30), "S-2": ("Y", 40)})  # S-1 absent but has a mark
    r = await client.post(f"/api/v1/excel/workbooks/{wid}/imports",
                          files={"file": ("wb.xlsx", filled, XLSX)},
                          headers={**h, "Idempotency-Key": uuid.uuid4().hex})
    errs = (await client.get(f"/api/v1/excel/imports/{r.json()['import_id']}/errors", headers=h)).json()["errors"]
    assert any(e["student_id"] == "S-1" and e["code"] == "ABSENT_STUDENT_HAS_MARKS" for e in errs)


# ---------- confirm applies once (idempotent) ----------

async def test_confirm_applies_once(client):
    h = await _admin(client)
    ctx = await _setup(client, h)
    wid, content = await _make_workbook_bytes(client, h, ctx)
    filled = _fill(content, {"S-1": ("Y", 60), "S-2": ("Y", 55)})
    imp = (await client.post(f"/api/v1/excel/workbooks/{wid}/imports",
                             files={"file": ("wb.xlsx", filled, XLSX)},
                             headers={**h, "Idempotency-Key": uuid.uuid4().hex})).json()
    assert imp["accepted_count"] == 2
    c1 = (await client.post(f"/api/v1/excel/imports/{imp['import_id']}/confirm", headers=h)).json()
    assert c1["rows_applied"] == 2 and c1["already"] is False
    # replay confirm -> no double apply
    c2 = (await client.post(f"/api/v1/excel/imports/{imp['import_id']}/confirm", headers=h)).json()
    assert c2["already"] is True
    # exactly two TotalMark rows exist for this exam
    from sqlalchemy import func, select
    from app.models.marks import TotalMark
    from app.models.exam import ExamStudentSubject, ExamStudent
    async with get_sessionmaker()() as db:
        cnt = await db.scalar(
            select(func.count()).select_from(TotalMark)
            .join(ExamStudentSubject, ExamStudentSubject.id == TotalMark.exam_student_subject_id)
            .join(ExamStudent, ExamStudent.id == ExamStudentSubject.exam_student_id)
            .where(ExamStudent.exam_id == ctx["exam_id"]))
    assert cnt == 2


# ---------- EXCEL writer enforcement ----------

async def test_confirm_rejected_when_scope_not_excel(client):
    h = await _admin(client)
    ctx = await _setup(client, h, writer="ONLINE")  # scope owned by ONLINE, not EXCEL
    wid, content = await _make_workbook_bytes(client, h, ctx)
    filled = _fill(content, {"S-1": ("Y", 60), "S-2": ("Y", 55)})
    imp = (await client.post(f"/api/v1/excel/workbooks/{wid}/imports",
                             files={"file": ("wb.xlsx", filled, XLSX)},
                             headers={**h, "Idempotency-Key": uuid.uuid4().hex})).json()
    r = await client.post(f"/api/v1/excel/imports/{imp['import_id']}/confirm", headers=h)
    assert r.status_code == 422 and r.json()["detail"]["code"] == "WRITER_MODE_MISMATCH"


async def test_reupload_same_key_is_idempotent(client):
    h = await _admin(client)
    ctx = await _setup(client, h)
    wid, content = await _make_workbook_bytes(client, h, ctx)
    filled = _fill(content, {"S-1": ("Y", 60), "S-2": ("Y", 55)})
    key = uuid.uuid4().hex
    imp1 = (await client.post(f"/api/v1/excel/workbooks/{wid}/imports",
                              files={"file": ("wb.xlsx", filled, XLSX)}, headers={**h, "Idempotency-Key": key})).json()
    imp2 = (await client.post(f"/api/v1/excel/workbooks/{wid}/imports",
                              files={"file": ("wb.xlsx", filled, XLSX)}, headers={**h, "Idempotency-Key": key})).json()
    assert imp1["import_id"] == imp2["import_id"]
